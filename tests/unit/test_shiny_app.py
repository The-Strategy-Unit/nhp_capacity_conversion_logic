import importlib.util
import os
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from types import ModuleType
from unittest.mock import call

import pandas as pd
import pytest
from openpyxl import load_workbook


def _load_app_module() -> ModuleType:
    app_path = Path(__file__).parents[2] / "app.py"
    spec = importlib.util.spec_from_file_location("capacity_conversion_app", app_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Unable to load {app_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


app = _load_app_module()

APP_ENVIRONMENT = {
    "AZ_STORAGE_EP": "https://storage.example.com",
    "AZ_STORAGE_RESULTS": "results",
    "AZ_TABLE_ENDPOINT": "https://table.example.com",
    "TABLE_NAME": "metadata",
}


def test_app_registers_favicon():
    assert app.FAVICON_DEPENDENCY.head is not None
    html = app.FAVICON_DEPENDENCY.head.get_html_string()

    assert '<link rel="icon" type="image/x-icon" href="favicon.ico"/>' in html
    assert app.app._static_assets["/"] == Path(app.STATIC_ASSETS_DIR)


def _functional_aggregation(**overrides) -> dict:
    entity = {
        "PartitionKey": "dev",
        "RowKey": "guid-123",
        "dataset": "RXX",
        "scenario_name": "scenario-a",
        "scenario_runtime": "20260817_143723",
    }
    entity.update(overrides)
    return entity


def test_catalogue_frame_validates_and_parses_entities():
    result = app._catalogue_frame([_functional_aggregation()])

    assert list(result.columns) == list(app.CATALOGUE_COLUMNS)
    assert result.loc[0, "scenario_runtime"] == pd.Timestamp("2026-08-17T14:37:23Z")


def test_catalogue_frame_handles_an_empty_catalogue():
    result = app._catalogue_frame([])

    assert result.empty
    assert list(result.columns) == list(app.CATALOGUE_COLUMNS)


def test_catalogue_frame_requires_all_columns():
    entity = _functional_aggregation()
    del entity["scenario_name"]

    with pytest.raises(
        ValueError,
        match="Catalogue is missing required columns: scenario_name",
    ):
        app._catalogue_frame([entity])


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("scenario_name", ""),
        ("scenario_runtime", "not-a-date"),
    ],
)
def test_catalogue_frame_ignores_invalid_entities(field, value, caplog):
    result = app._catalogue_frame([_functional_aggregation(**{field: value})])

    assert result.empty
    assert "Ignored 1 catalogue entities" in caplog.text


def test_catalogue_frame_preserves_valid_entities_when_another_is_invalid():
    invalid = _functional_aggregation(RowKey="invalid-guid")
    del invalid["scenario_name"]

    result = app._catalogue_frame([_functional_aggregation(), invalid])

    assert result["RowKey"].tolist() == ["guid-123"]


def test_catalogue_frame_ignores_an_inconsistent_partition_key():
    result = app._catalogue_frame([_functional_aggregation(PartitionKey="prod")])

    assert result.empty


def _functional_aggregation_catalogue() -> pd.DataFrame:
    return app._catalogue_frame(
        [
            _functional_aggregation(),
            _functional_aggregation(
                RowKey="guid-other-dataset",
                dataset="RYY",
            ),
        ]
    )


def test_filter_functional_aggregations_applies_provider_entitlement():
    result = app._filter_functional_aggregations_for_user(
        _functional_aggregation_catalogue(),
        ["nhp_provider_RXX", "unrelated_group"],
        is_local=False,
    )

    assert result["RowKey"].tolist() == ["guid-123"]


@pytest.mark.parametrize("group", ["nhp_devs", "nhp_power_users"])
def test_filter_functional_aggregations_allows_privileged_groups(group):
    result = app._filter_functional_aggregations_for_user(
        _functional_aggregation_catalogue(),
        [group],
        is_local=False,
    )

    assert result["RowKey"].tolist() == [
        "guid-123",
        "guid-other-dataset",
    ]


def test_filter_functional_aggregations_allows_all_runs_locally():
    result = app._filter_functional_aggregations_for_user(
        _functional_aggregation_catalogue(),
        None,
        is_local=True,
    )

    assert result["RowKey"].tolist() == [
        "guid-123",
        "guid-other-dataset",
    ]


def test_filter_functional_aggregations_fails_closed_without_connect_groups():
    result = app._filter_functional_aggregations_for_user(
        _functional_aggregation_catalogue(),
        None,
        is_local=False,
    )

    assert result.empty


def test_functional_aggregation_choices_are_newest_first():
    functional_aggregations = app._catalogue_frame(
        [
            _functional_aggregation(),
            _functional_aggregation(
                RowKey="guid-newer",
                scenario_runtime="20260818_090500",
            ),
        ]
    )

    result = app._functional_aggregation_choices(functional_aggregations)

    assert result == {
        "guid-newer": "18 Aug 2026, 09:05 UTC",
        "guid-123": "17 Aug 2026, 14:37 UTC",
    }


def test_authorise_functional_aggregation_revalidates_the_selected_entity():
    entity = _functional_aggregation()

    result = app._authorise_functional_aggregation(
        entity,
        dataset="RXX",
        scenario="scenario-a",
        functional_aggregation_guid="guid-123",
        groups=["nhp_provider_RXX"],
        is_local=False,
    )

    assert result == entity
    assert result is not entity


@pytest.mark.parametrize(
    ("selection", "groups"),
    [
        ({"dataset": "RYY"}, ["nhp_provider_RXX"]),
        ({"scenario": "different-scenario"}, ["nhp_provider_RXX"]),
        (
            {"functional_aggregation_guid": "different-guid"},
            ["nhp_provider_RXX"],
        ),
        ({}, ["nhp_provider_RYY"]),
    ],
)
def test_authorise_functional_aggregation_rejects_stale_or_unauthorised_selections(
    selection,
    groups,
):
    expected_selection = {
        "dataset": "RXX",
        "scenario": "scenario-a",
        "functional_aggregation_guid": "guid-123",
    } | selection

    with pytest.raises(PermissionError, match="not available"):
        app._authorise_functional_aggregation(
            _functional_aggregation(),
            **expected_selection,
            groups=groups,
            is_local=False,
        )


def test_is_local_development(mocker):
    mocker.patch.dict(os.environ, {}, clear=True)
    assert app._is_local_development()

    mocker.patch.dict(os.environ, {"POSIT_PRODUCT": "CONNECT"})
    assert not app._is_local_development()


def test_load_capacity_results(mocker):
    mocker.patch.dict(os.environ, APP_ENVIRONMENT, clear=True)
    functional_aggregation = _functional_aggregation(
        Timestamp=datetime(2026, 8, 17, 14, 50, tzinfo=UTC),
    )
    create_path = mocker.patch.object(
        app,
        "create_aggregations_path",
        return_value="functional-aggregations/dev/guid-123/",
    )
    aggregations = pd.DataFrame({"total": [1]})
    load_aggregation = mocker.patch.object(
        app,
        "load_aggregations",
        return_value=aggregations,
    )
    filter_aggregation = mocker.patch.object(
        app,
        "filter_aggregations",
        return_value=aggregations,
    )
    assumptions = pd.DataFrame({"Value": [1]})
    mocker.patch.object(app, "load_assumptions", return_value=assumptions)
    process = mocker.patch.object(app, "process_activity_type")

    data_to_save = app._load_capacity_results(functional_aggregation)

    metadata = create_path.call_args.args[0]
    assert metadata["guid"] == "guid-123"
    assert metadata["capacity_model_version"] == "dev"
    assert metadata["Timestamp"] == "2026-08-17T14:50:00+00:00"
    create_path.assert_called_once_with(metadata)
    load_aggregation.assert_has_calls(
        [
            call(
                "https://storage.example.com",
                "results",
                "functional-aggregations/dev/guid-123/",
                activity_type,
            )
            for activity_type in app.ACTIVITY_TYPES
        ]
    )
    assert load_aggregation.call_count == 6
    assert filter_aggregation.call_count == 6
    process.assert_has_calls(
        [
            call(
                "op",
                aggregations,
                app.calculate_op_capacity,
                assumptions,
                data_to_save,
                preprocess=None,
            ),
            call(
                "aae",
                aggregations,
                app.calculate_aae_capacity,
                assumptions,
                data_to_save,
                preprocess=None,
            ),
            call(
                "ip_daycase",
                aggregations,
                app.calculate_daycase_capacity,
                assumptions,
                data_to_save,
                preprocess=None,
            ),
            call(
                "ip_maternity",
                aggregations,
                app.calculate_maternity_capacity,
                assumptions,
                data_to_save,
                preprocess=app.preprocess_ip_maternity_data,
            ),
            call(
                "ip_wards",
                aggregations,
                app.calculate_ip_wards_capacity,
                assumptions,
                data_to_save,
                preprocess=app.preprocess_ip_wards_data,
            ),
            call(
                "ip_procedures_and_theatres",
                aggregations,
                app.calculate_ip_theatres_capacity,
                assumptions,
                data_to_save,
                preprocess=app.preprocess_ip_theatres_data,
            ),
        ]
    )
    assert process.call_count == 6
    runtime = data_to_save["metadata"].loc["capacity_conversion_runtime"]
    assert len(runtime) == 15
    assert runtime[8] == "_"
    assert runtime.replace("_", "").isdigit()
    assert data_to_save["metadata"].loc[list(app.SITES)].to_dict() == app.SITES


def test_load_capacity_results_requires_storage_configuration(mocker):
    mocker.patch.dict(
        os.environ,
        {
            key: value
            for key, value in APP_ENVIRONMENT.items()
            if key != "AZ_STORAGE_RESULTS"
        },
        clear=True,
    )

    with pytest.raises(
        RuntimeError,
        match="Missing required environment variable: AZ_STORAGE_RESULTS",
    ):
        app._load_capacity_results(_functional_aggregation())


def test_create_workbook():
    capacity = pd.DataFrame(
        {
            "output": ["room", "room"],
            "model_run": [1, 2],
            "value": [1.0, 3.0],
        }
    ).set_index(["output", "model_run"])
    data_to_save = {
        "metadata": pd.Series({"guid": "guid-123"}),
        "op_capacity": capacity,
    }

    workbook = load_workbook(BytesIO(app._create_workbook(data_to_save)))

    assert workbook.sheetnames == ["metadata", "op_capacity"]
    rows = list(workbook["op_capacity"].values)
    assert rows[0] == ("output", "p10", "mean", "p90")
    assert rows[1] == ("room", 1.2, 2, 2.8)

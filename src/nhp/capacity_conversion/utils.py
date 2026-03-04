import pandas as pd
from nhpy.utils import get_logger
from azure.data.tables import TableClient
from azure.identity import DefaultAzureCredential
from azure.core.exceptions import ResourceNotFoundError
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import os

logger = get_logger()


def calculate_prediction_intervals_and_mean(
    activity_column: pd.Series,
) -> dict[str, float]:
    """Calculate p10, p90 and mean for activity in each functional area

    Args:
        activity_column (pd.Series): Column with activity counts for each functional area

    Returns:
        dict[str, float]: Dictionary with p10, p90 and mean as keys
    """
    results_dict = {"mean": float(activity_column.mean())}
    results_dict["p10"] = float(activity_column.quantile(0.1))
    results_dict["p90"] = float(activity_column.quantile(0.9))
    return results_dict


def load_assumptions(path_to_csv: str) -> pd.DataFrame:
    """Loads assumptions for use in model
    TODO: #7 Currently loads from CSV but later we should allow users to set in another way

    Args:
        path_to_csv (str): Path to assumptions csv file

    Returns:
        pd.DataFrame: Dataframe with assumption values and variable names
    """
    logger.info(f"Loading assumptions from {path_to_csv}...")
    return pd.read_csv(path_to_csv).set_index("assumption_name").sort_index()


def save_results_to_excel(data_to_save: dict[str, pd.DataFrame | pd.Series]) -> None:
    """Saves results of capacity conversion pipeline to Excel

    Args:
        data_to_save (dict[str, pd.DataFrame  |  pd.Series]): Dictionary of data to save, where the keys are the titles of the
        worksheets and the values are the dataframes to be included. At minimum should include "metadata" key and dataframe.
    """
    directory = os.path.join(
        "results",
        str(data_to_save["metadata"].loc["guid"]),
        str(data_to_save["metadata"].loc["capacity_conversion_runtime"]),
    )
    os.makedirs(directory, exist_ok=True)
    filepath = os.path.join(directory, "capacity_conversion_results.xlsx")
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    for sheet_name, df in data_to_save.items():
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(
            dataframe_to_rows(pd.DataFrame(df).reset_index(), index=False, header=True),
            start=1,
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(filepath)
    logger.info(f"💾 Results saved to {filepath}")


def load_metadata_from_ats(
    guid: str,
    storage_endpoint: str,
    table_name: str,
    capacity_model_version: str,
) -> dict:
    """Loads metadata for scenario converted to functional area aggregations
    from Azure Table Storage

    Args:
        guid (str): GUID for functional area aggregation
        storage_endpoint (str): Azure Table Storage endpoint, in format "https://{storage_account_name}.table.core.windows.net"
        table_name (str): Table name containing metadata for Functional Area Aggregations
        capacity_model_version (str): Version of capacity model.

    Returns:
        dict: Dictionary with metadata for given Functional Area aggregation
    """
    credential = DefaultAzureCredential()
    table_client = TableClient(
        endpoint=storage_endpoint, table_name=table_name, credential=credential
    )
    try:
        entity = table_client.get_entity(
            partition_key=capacity_model_version, row_key=guid
        )
        metadata = dict(entity)
        metadata["guid"] = guid
        metadata["capacity_model_version"] = capacity_model_version
        return metadata
    except ResourceNotFoundError:
        raise


def create_aggregations_path(metadata: dict) -> str:
    """Create path to aggregations parquet files on Azure Storage

    Args:
        metadata (dict): Dictionary of metadata for capacity conversion

    Returns:
        str: Full path to the specific functional area aggregations to be converted to capacity
    """
    return f"functional-aggregations/{metadata['capacity_model_version']}/{metadata['guid']}/"


def validate_required_env_vars() -> dict:
    """
    Loads environment variables and ensures required variables are present.
    Raises EnvironmentError if any are missing or empty.
    Returns a dictionary of the validated variables.
    """

    load_dotenv()

    required_vars = [
        "AZ_STORAGE_EP",
        "AZ_STORAGE_RESULTS",
        "TABLE_NAME",
        "AZ_TABLE_ENDPOINT",
    ]

    values = {}
    missing = []

    for var in required_vars:
        value = os.getenv(var)
        if not value:
            missing.append(var)
        else:
            values[var] = value

    if missing:
        raise EnvironmentError(
            f"Missing required environment variables in .env: {', '.join(missing)}"
        )

    return values

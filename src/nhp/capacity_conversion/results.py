import logging
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


def tidy_metadata(
    data_to_save: dict[str, pd.DataFrame | pd.Series],
) -> dict[str, pd.DataFrame | pd.Series]:
    if "metadata" in data_to_save:
        rename = {
            "app_version": "demand_model_version",
            "scenario_name": "demand_model_scenario_name",
            "scenario_runtime": "demand_model_scenario_runtime",
        }

        keep = [
            "dataset",
            "capacity_model_version",
            "ip_sites",
            "op_sites",
            "aae_sites",
            "capacity_conversion_runtime",
            *rename,
        ]

        metadata = data_to_save["metadata"]

        data_to_save["metadata"] = metadata.reindex(
            [key for key in keep if key in metadata.index]
        ).rename(index=rename)
    return data_to_save


def summarise_model_runs(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate p10, p90 and mean across all model runs

    Args:
        df (pd.DataFrame): MultiIndex DataFrame with one index called "model_run"

    Raises:
        ValueError: If more than one value column in DataFrame

    Returns:
        pd.DataFrame: Summarised DataFrame
    """
    group_col_names = [name for name in df.index.names if name != "model_run"]
    if len(group_col_names) != 1:
        raise ValueError("Expected exactly one index column.")
    value_cols = [c for c in df.columns if c != "model_run"]
    if len(value_cols) > 1:
        df_list = []
        for col in value_cols:
            summary_df = pd.DataFrame(
                df.groupby(level=group_col_names)[col].agg(
                    p10=lambda s: s.quantile(0.10),
                    mean="mean",
                    p90=lambda s: s.quantile(0.90),
                )
            )
            summary_df["measure"] = col
            df_list.append(summary_df.reset_index())
        return pd.concat(df_list).set_index(group_col_names + ["measure"]).sort_index()
    return pd.DataFrame(
        df.groupby(level=group_col_names)[value_cols[0]].agg(
            p10=lambda s: s.quantile(0.10),
            mean="mean",
            p90=lambda s: s.quantile(0.90),
        )
    )


def process_and_save_results_to_excel(
    data_to_save: dict[str, pd.DataFrame | pd.Series],
) -> None:
    """Saves results of capacity conversion pipeline to Excel

    Args:
        data_to_save (dict[str, pd.DataFrame  |  pd.Series]): Dictionary of data to save, where the keys are the titles of the
        worksheets and the values are the dataframes to be included. At minimum should include "metadata" key and dataframe.
    """
    directory = os.path.join(
        "results",
        str(data_to_save["metadata"].loc["guid"]),
        str(data_to_save["metadata"].loc["capacity_conversion_runtime"]),
    )
    os.makedirs(directory, exist_ok=True)
    filepath = os.path.join(directory, "capacity_conversion_results.xlsx")
    wb = Workbook()
    default_sheet = wb.active
    assert default_sheet is not None
    wb.remove(default_sheet)
    data_to_save = tidy_metadata(data_to_save)
    for sheet_name, df in data_to_save.items():
        if isinstance(df, pd.DataFrame) and "model_run" in df.index.names:
            df = summarise_model_runs(df)
        ws = wb.create_sheet(title=sheet_name[:31])
        if isinstance(df, pd.Series):
            rows = dataframe_to_rows(
                df.to_frame().reset_index(),
                index=False,
                header=False,
            )
        else:
            rows = dataframe_to_rows(
                df.reset_index(),
                index=False,
                header=True,
            )

        for r_idx, row in enumerate(
            rows,
            start=1,
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
    wb.save(filepath)
    logger.info(f"💾 Results saved to {filepath}")
